import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import warnings  # For cleanly catching and suppressing external workbook style warnings

# Tkinter modules for the UI
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# Third-party module for native drag and drop capabilities
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
except ImportError:
    pass

# ----------------------------------------------------
# GLOBAL EXPLICIT VARIABLES INITIALIZATION
# ----------------------------------------------------
file_layout_dict_b01 = {}
file_layout_dict_b02 = {}
file_layout_dict_p01 = {}

# ==========================================
# CONFIGURATION LOADER ENGINE
# ==========================================
def get_resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def load_layout_configs():
    """Loads parsing schemas from local external JSON configs."""
    configs = {}
    files = {
        "B01": "configs/b01_config.json",
        "B02": "configs/b02_config.json",
        "P01": "configs/p01_config.json"
    }
    
    for key, filename in files.items():
        resolved_path = get_resource_path(filename)
        if not os.path.exists(resolved_path):
            raise FileNotFoundError(
                f"Missing critical layout specification schema: '{filename}'.\n"
                f"Resolved destination attempted: {resolved_path}"
            )
        with open(resolved_path, "r", encoding="utf-8") as f:
            configs[key] = json.load(f)
            
    return configs["B01"], configs["B02"], configs["P01"]

try:
    file_layout_dict_b01, file_layout_dict_b02, file_layout_dict_p01 = load_layout_configs()
except Exception as err:
    root_err = tk.Tk()
    root_err.withdraw()
    messagebox.showerror("Initialization Error", str(err))
    exit(1)


# ==========================================
# PARSING ENGINE LOGIC
# ==========================================
def retrieve_specific_entry(line, layout_dict, entry_name):
    if entry_name not in layout_dict:
        raise ValueError(f"Entry name '{entry_name}' not found in layout dictionary.")
    offset = layout_dict[entry_name]["offset"]
    length = layout_dict[entry_name]["length"]
    val = line[offset:offset + length].strip()
    
    if val.replace('.', '', 1).isdigit():
        return float(val) if '.' in val else int(val)
    return val

def retrieve_all_entries(line, layout_dict):
    return {entry_name: retrieve_specific_entry(line, layout_dict, entry_name) for entry_name in layout_dict}

def parse_bk_file(filepath):
    products = []
    current_product = {}
    current_product_name = None

    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.rstrip('\n')
            if not line:
                continue
            
            record_type = line[:3]
            
            if record_type == "B01":
                if current_product and current_product_name:
                    products.append(current_product)
                
                offset = file_layout_dict_b01["product_description"]["offset"]
                length = file_layout_dict_b01["product_description"]["length"]
                product_name = line[offset:offset + length].strip()

                current_product_name = product_name
                current_product = {
                    current_product_name: {
                        "B01": line, "B02": None, "P01": None
                    }
                }
            elif record_type == "B02" and current_product_name:
                current_product[current_product_name]["B02"] = line
            elif record_type == "P01" and current_product_name:
                current_product[current_product_name]["P01"] = line

        if current_product and current_product_name:
            products.append(current_product)

    for product in products:
        name = list(product.keys())[0]
        product[name]["B01_entries"] = retrieve_all_entries(product[name]["B01"] or " "*210, file_layout_dict_b01)
        product[name]["B02_entries"] = retrieve_all_entries(product[name]["B02"] or " "*210, file_layout_dict_b02)
        product[name]["P01_entries"] = retrieve_all_entries(product[name]["P01"] or " "*240, file_layout_dict_p01)
        
    return products

# ==========================================
# Export to Excel logic
# ==========================================
def export_to_excel(products, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TGP POS Translations"
    ws.views.sheetView[0].showGridLines = True
    
    font_section_header = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
    font_sub_header = Font(name="Segoe UI", size=10, bold=True, color="333333")
    font_data = Font(name="Segoe UI", size=10, color="222222")
    
    navy_header_fill = PatternFill(start_color="2F3E46", end_color="2F3E46", fill_type="solid")
    b01_sub_fill = PatternFill(start_color="DCEEFF", end_color="DCEEFF", fill_type="solid")
    b01_row_fill_a = PatternFill(start_color="F2F8FF", end_color="F2F8FF", fill_type="solid")
    b01_row_fill_b = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    
    b02_sub_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid") 
    b02_row_fill_a = PatternFill(start_color="F4F9F1", end_color="F4F9F1", fill_type="solid") 
    b02_row_fill_b = PatternFill(start_color="EBF5E6", end_color="EBF5E6", fill_type="solid") 
    
    p01_sub_fill = PatternFill(start_color="E8E1F5", end_color="E8E1F5", fill_type="solid") 
    p01_row_fill_a = PatternFill(start_color="F6F3FA", end_color="F6F3FA", fill_type="solid") 
    p01_row_fill_b = PatternFill(start_color="EFEAF6", end_color="EFEAF6", fill_type="solid") 

    thin_border_side = Side(style='thin', color='D9D9D9')
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    segments_config = [
        {"name": "B01", "layout": file_layout_dict_b01, "sub_fill": b01_sub_fill, "fill_a": b01_row_fill_a, "fill_b": b01_row_fill_b, "entry_key": "B01_entries"},
        {"name": "B02", "layout": file_layout_dict_b02, "sub_fill": b02_sub_fill, "fill_a": b02_row_fill_a, "fill_b": b02_row_fill_b, "entry_key": "B02_entries"},
        {"name": "P01", "layout": file_layout_dict_p01, "sub_fill": p01_sub_fill, "fill_a": p01_row_fill_a, "fill_b": p01_row_fill_b, "entry_key": "P01_entries"}
    ]
    
    current_col = 1
    for segment in segments_config:
        fields = list(segment["layout"].keys())
        num_fields = len(fields)
        if num_fields == 0:
            continue
            
        start_col = current_col
        end_col = current_col + num_fields - 1
        
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        header_cell = ws.cell(row=1, column=start_col, value=segment["name"])
        header_cell.font = font_section_header
        header_cell.fill = navy_header_fill
        header_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        
        for field_idx, field_key in enumerate(fields):
            col_pos = start_col + field_idx
            clean_title = field_key.replace("_", " ").title()
            
            sub_cell = ws.cell(row=2, column=col_pos, value=clean_title)
            sub_cell.font = font_sub_header
            sub_cell.fill = segment["sub_fill"]
            sub_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            sub_cell.border = cell_border
            
        current_col = end_col + 1

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 26
    
    for idx, product in enumerate(products):
        current_row = idx + 3
        prod_key = list(product.keys())[0]
        data_map = product[prod_key]
        is_zebra_row = (idx % 2 == 0)
        segment_col_start = 1
        
        for segment in segments_config:
            fields = list(segment["layout"].keys())
            if not fields:
                continue
            current_fill = segment["fill_b"] if is_zebra_row else segment["fill_a"]
            
            for field_idx, field_key in enumerate(fields):
                col_pos = segment_col_start + field_idx
                val = data_map[segment["entry_key"]].get(field_key, "")
                
                cell = ws.cell(row=current_row, column=col_pos, value=val)
                cell.font = font_data
                cell.fill = current_fill
                cell.border = cell_border
                
                if isinstance(val, float):
                    cell.number_format = "$#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif isinstance(val, int):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
            segment_col_start += len(fields)
        ws.row_dimensions[current_row].height = 20

    ws.freeze_panes = "A3" 
    
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value and len(str(cell.value)) > max_len:
                max_len = len(str(cell.value))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
        
    wb.save(filename)
    

# ==========================================
# CORE UI WINDOW CLASS
# ==========================================
class BK1ConverterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("BK1 Flat-File Parser")
        self.root.geometry("540x550") 
        self.root.resizable(True, True)
        
        self.bg_base = "#1F232A"
        self.bg_surface = "#2D3139"
        self.text_primary = "#E2E8F0"
        self.text_secondary = "#8E8E93" 
        self.ios_blue = "#0A84FF"      
        self.ios_green = "#30D158"      
        self.ios_red = "#FF453A"        
        
        self.root.configure(bg=self.bg_base)
        
        try:
            from ctypes import windll, byref, c_int, sizeof
            HWND = windll.user32.GetParent(self.root.winfo_id())
            windll.dwmapi.DwmSetWindowAttribute(HWND, 20, byref(c_int(1)), sizeof(c_int))
        except Exception:
            pass

        self.parsed_data = None
        self.loaded_filename = ""
        self.font_family = ("-apple-system", "SF Pro Text", "Helvetica Neue", "Segoe UI", "Arial")

        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        self.style.configure("TFrame", background=self.bg_base)
        self.style.configure("Surface.TFrame", background=self.bg_surface)
        self.style.configure("TLabel", background=self.bg_base, foreground=self.text_primary)
        self.style.configure("Surface.TLabel", background=self.bg_surface, foreground=self.text_primary)
        
        self.style.configure(
            "IOS.TButton", 
            background=self.bg_surface, 
            foreground=self.ios_blue, 
            font=(self.font_family, 10, "bold"),
            borderwidth=1,
            bordercolor=self.ios_blue,
            padding=(10, 4)
        )
        self.style.map(
            "IOS.TButton", 
            background=[("active", "#3A3F4B")],
            foreground=[("active", "#64B5FF")]
        )
        
        # --- FIXED SPLIT SCREEN CONTAINER LAYOUT ---
        # Top Frame: Handles variable items (grows/scrolls organically)
        self.top_scrollable_container = ttk.Frame(root, padding="20")
        self.top_scrollable_container.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        # Bottom Frame: Rigid control panel completely locked to the base
        self.bottom_fixed_container = ttk.Frame(root, padding="20")
        self.bottom_fixed_container.pack(side=tk.BOTTOM, fill=tk.X)

        # Populate Top Frame Elements
        title_lbl = ttk.Label(self.top_scrollable_container, text="TGS POS File Layout Converter", font=(self.font_family, 14, "bold"))
        title_lbl.pack(pady=(0, 10), anchor="w")

        self.drop_canvas = tk.Canvas(self.top_scrollable_container, bg=self.bg_surface, highlightthickness=0, height=100)
        self.drop_canvas.pack(fill=tk.X, pady=5)
        self.drop_canvas.create_rectangle(4, 4, 494, 96, outline="#48484A", dash=(4, 4), width=2)
        
        self.file_label = tk.Label(
            self.drop_canvas, 
            text="Drag & Drop .bk File Here", 
            font=(self.font_family, 10, "bold"), 
            bg=self.bg_surface,
            fg=self.text_primary,
            justify="center"
        )
        self.drop_canvas.create_window(250, 25, window=self.file_label, anchor="center")
        
        self.or_lbl = tk.Label(self.drop_canvas, text="— or —", font=(self.font_family, 9), bg=self.bg_surface, fg=self.text_secondary)
        self.drop_canvas.create_window(250, 50, window=self.or_lbl, anchor="center")

        self.upload_btn = ttk.Button(self.drop_canvas, text="Browse Files", style="IOS.TButton", command=self.load_file)
        self.drop_canvas.create_window(250, 75, window=self.upload_btn, anchor="center")

        self.status_lbl = ttk.Label(
            self.top_scrollable_container, 
            text="Waiting for .bk file...", 
            font=(self.font_family, 10), 
            foreground=self.text_secondary,
            wraplength=440,
            justify="center"
        )
        self.status_lbl.pack(pady=5, anchor="center")

        self.save_btn = tk.Button(
            self.top_scrollable_container, 
            text="Export to Excel", 
            font=(self.font_family, 10, "bold"),
            bg="#2C2C2E", fg="#48484A", state=tk.DISABLED, relief="flat",
            borderwidth=0, highlightthickness=0, bd=0, command=self.save_file, cursor="arrow"
        )
        self.save_btn.pack(fill=tk.X, ipady=6, pady=2)
        
        self.upload_files_to_search_btn = tk.Button(
            self.top_scrollable_container,
            text="Upload excel POS files to search",
            font=(self.font_family, 10, "bold"),
            bg="#9BFF97", fg="#111111", relief="flat",
            borderwidth=0, highlightthickness=0, bd=0, command=self.upload_files_to_search, cursor="arrow"
        )
        self.upload_files_to_search_btn.pack(fill=tk.X, ipady=6, pady=2)

        # Added a scrollable frame for uploaded file lists so they never overflow down over the search components
        list_scroll_frame = ttk.Frame(self.top_scrollable_container)
        list_scroll_frame.pack(fill=tk.BOTH, expand=True, pady=4)

        self.search_files_lbl = ttk.Label(
            list_scroll_frame,
            text="",
            font=(self.font_family, 9),
            foreground=self.text_secondary,
            wraplength=460,
            justify="left"
        )
        self.search_files_lbl.pack(fill=tk.BOTH, expand=True, anchor="w")

        # Populate Bottom Fixed Frame Elements (Locked in Position)
        self.search_controls_frame = ttk.Frame(self.bottom_fixed_container)
        self.search_entry = tk.Entry(
            self.search_controls_frame,
            font=(self.font_family, 12),  # Expanded text readability size
            relief="solid", bd=1, bg=self.bg_surface, fg=self.text_primary, insertbackground="white"
        )
        self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=6, ipadx=6)

        self.search_btn = tk.Button(
            self.search_controls_frame,
            text="Search",
            font=(self.font_family, 10, "bold"),
            bg="#9BFF97", fg="#111111", relief="flat",
            borderwidth=0, highlightthickness=0, bd=0,
            state=tk.DISABLED, command=self.search_uploaded_files, cursor="arrow"
        )
        self.search_btn.pack(side=tk.LEFT, padx=(6, 0), ipady=6, ipadx=12)
        self.search_controls_frame.pack(fill=tk.X, pady=(0, 5))  # Always visible context structure

        self.result_text_area = tk.Text(
            self.bottom_fixed_container,
            font=(self.font_family, 9),
            bg=self.bg_base, fg=self.text_secondary,
            bd=0, highlightthickness=0, height=5, state=tk.DISABLED
        )
        self.result_text_area.pack(fill=tk.X, pady=(5, 0))

        self.uploaded_search_files = []
        self.uploaded_search_workbooks = []

        self.setup_drag_and_drop()
    
    def load_file(self):
        file_path = filedialog.askopenfilename(
            title="Select product catalog file",
            filetypes=[("BK files", "*.bk*"), ("All files", "*.*")]
        )
        if file_path:
            self.process_and_validate_file(file_path)

    def process_and_validate_file(self, file_path):
        if not os.path.splitext(os.path.basename(file_path).lower())[1].startswith(".bk"):
            self.status_lbl.config(text="Uploaded file was not a .bk# file", foreground=self.ios_red)
            return

        try:
            self.status_lbl.config(text="Processing .bk file (might take a moment if large)...", foreground=self.ios_blue)
            self.root.update_idletasks()
            
            self.parsed_data = parse_bk_file(file_path)
            self.loaded_filename = os.path.basename(file_path)
            
            self.file_label.config(text=f"Loaded: {self.loaded_filename}", fg=self.ios_green)
            self.status_lbl.config(text=f"Success! Found {len(self.parsed_data)} POS items.", foreground=self.ios_green) 
            
            self.save_btn.config(
                state=tk.NORMAL, bg=self.ios_blue, fg="white", 
                activebackground="#0070E6", activeforeground="white", cursor="hand2"
            )
        except Exception as e:
            self.status_lbl.config(text=f"Parsing failure: {str(e)}", foreground=self.ios_red) 

    def save_file(self):
        if not self.parsed_data:
            return
        
        out_path = filedialog.asksaveasfilename(
            title="Export Transpiled Dataset",
            initialfile=os.path.splitext(self.loaded_filename)[0],
            filetypes=[("Excel Spreadsheet", "*.xlsx"), ("JSON Matrix Document", "*.json")]
        )
        if not out_path:
            return

        try:
            self.status_lbl.config(text="Writing to excel (might take a moment if large)...", foreground=self.ios_blue)
            self.root.update_idletasks()

            if out_path.endswith(".json"):
                with open(out_path, "w", encoding="utf-8") as f:
                    json.dump(self.parsed_data, f, indent=4)
            else:
                if not out_path.endswith(".xlsx"):
                    out_path += ".xlsx"
                export_to_excel(self.parsed_data, out_path)
                
            self.status_lbl.config(
                text=f"Excel file saved successfully to: {os.path.basename(out_path)}", 
                foreground=self.ios_green
            )
        except Exception as e:
            error_msg = str(e)
            if "Permission denied" in error_msg:
                error_msg = "Permission Denied. Close the file if it's open in Excel and retry."
            
            self.status_lbl.config(text=f"Export failed: {error_msg}", foreground=self.ios_red)

    def upload_files_to_search(self):
        file_paths = filedialog.askopenfilenames(
            title="Select Excel POS files to search",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_paths:
            # 1. IMMEDIATE LOADING FEEDBACK UPDATES
            self.search_files_lbl.config(text="⏳ Uploading please wait...", foreground=self.ios_blue)
            self.search_btn.config(state=tk.DISABLED)
            self.root.update()  # Forces Tkinter to redraw the UI immediately

            self.uploaded_search_files = list(file_paths)
            self.uploaded_search_workbooks = []
            success_files = []
            failed_files = []

            for path in self.uploaded_search_files:
                try:
                    with warnings.catch_warnings():
                        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
                        wb = openpyxl.load_workbook(path, data_only=True)
                    
                    self.uploaded_search_workbooks.append((path, wb))
                    success_files.append(os.path.basename(path))
                except Exception:
                    failed_files.append(os.path.basename(path))

            # 2. RESTORE STATUS DISPLAYS ONCE COMPLETE
            if success_files:
                loaded_text = "🎯 Uploaded Target Files:\n" + "\n".join([f" • {f}" for f in success_files])
                self.search_files_lbl.config(text=loaded_text, foreground=self.text_primary)
                self.search_btn.config(state=tk.NORMAL)
                
                self.result_text_area.config(state=tk.NORMAL)
                self.result_text_area.delete("1.0", tk.END)
                self.result_text_area.insert(tk.END, "Enter a value to search the uploaded files for exact matches.")
                self.result_text_area.config(state=tk.DISABLED, fg=self.text_secondary)
                
                self.search_entry.delete(0, tk.END)
                self.search_entry.focus_set()

            if failed_files:
                messagebox.showwarning(
                    "Load Error",
                    f"Could not load the following Excel files:\n" + "\n".join(failed_files)
                )

    def search_uploaded_files(self):
        query = self.search_entry.get().strip()
        
        self.result_text_area.config(state=tk.NORMAL)
        self.result_text_area.delete("1.0", tk.END)
        
        if not query:
            self.result_text_area.insert(tk.END, "⚠️ Please type a target verification value.")
            self.result_text_area.config(state=tk.DISABLED, fg=self.ios_red)
            return

        # 1. IMMEDIATE SEARCHING FEEDBACK UPDATES
        self.result_text_area.insert(tk.END, "🔍 Searching please wait...")
        self.result_text_area.config(fg=self.ios_blue)
        self.root.update()  # Forces Tkinter to redraw text block before iterating files

        matches = []
        
        for path, workbook in self.uploaded_search_workbooks:
            filename = os.path.basename(path)
            file_has_matches = False
            file_matches = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    for cell_value in row:
                        if cell_value is None:
                            continue
                        
                        if str(cell_value).strip() == query:
                            file_matches.append(f"   ↳ Row {row_idx} (Sheet: {sheet_name})")
                            file_has_matches = True
                            
            # If this specific file had hits, add its header and its corresponding rows
            if file_has_matches:
                matches.append(f"📁 {filename}:")
                matches.extend(file_matches)

        # 2. CLEAR SEARCH STATUS AND SHOW RESULTS
        self.result_text_area.delete("1.0", tk.END)

        if matches:
            output_report = "✅ Matches Discovered:\n" + "\n".join(matches)
            self.result_text_area.insert(tk.END, output_report)
            self.result_text_area.config(state=tk.DISABLED, fg=self.ios_green)
        else:
            self.result_text_area.insert(tk.END, f"❌ Exact value '{query}' not found anywhere in selected matrices.")
            self.result_text_area.config(state=tk.DISABLED, fg=self.ios_red)
    def setup_drag_and_drop(self):
        try:
            self.root.drop_target_register(DND_FILES)
            self.root.dnd_bind('<<Drop>>', self.handle_dropped_file)
            
            self.drop_canvas.drop_target_register(DND_FILES)
            self.drop_canvas.dnd_bind('<<Drop>>', self.handle_dropped_file)
        except NameError:
            pass

    def handle_dropped_file(self, event):
        file_path = event.data
        if file_path.startswith('{') and file_path.endswith('}'):
            file_path = file_path[1:-1]
        if file_path:
            self.process_and_validate_file(file_path)


if __name__ == "__main__":
    try:
        window = TkinterDnD.Tk()
    except NameError:
        window = tk.Tk()
        
    app = BK1ConverterApp(window)
    window.mainloop()