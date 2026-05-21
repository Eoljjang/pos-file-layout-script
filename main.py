BK1_FILEPATH = "./sample/aginprod.bk1"
EXCEL_OUTPUT_PATH = "./output/product_catalog_master.xlsx"
import json 
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# This dict holds the mappings for each product number for B01
file_layout_dict_b01 = {
    "type": {"offset": 0, "length": 1},
    "code": {"offset": 1, "length": 2},
    "product_number": {"offset": 3, "length": 10},
    "product_description": {"offset": 13, "length": 35},
    "pack": {"offset": 48, "length": 4},
    "size": {"offset": 52, "length": 4},
    "measure": {"offset": 56, "length": 2},
    "scan_description": {"offset": 58, "length": 12},
    "retail_by_weight": {"offset": 70, "length": 1},
    "charge_tax_1_pst_hst": {"offset": 71, "length": 1},
    "charge_tax_2_unit_gst": {"offset": 72, "length": 1},
    "charge_tax_3_case_gst": {"offset": 73, "length": 1},
    "bottle_deposit_amount": {"offset": 74, "length": 4},
    "item_status": {"offset": 78, "length": 1},
    "dsd_product": {"offset": 79, "length": 1},
    "food_service_or_retail": {"offset": 80, "length": 1},
    "stock_status_id": {"offset": 81, "length": 1},
    "restriction_01_prescription_drugs": {"offset": 82, "length": 1},
    "restriction_02_YTTobTax": {"offset": 83, "length": 1},
    "restriction_03_SKHBC": {"offset": 84, "length": 1},
    "restriction_04_HdwCoop": {"offset": 85, "length": 1},
    "restriction_05_HdwSelCoop": {"offset": 86, "length": 1},
    "restriction_06_BCTobTax": {"offset": 87, "length": 1},
    "restriction_07_BCHBC": {"offset": 88, "length": 1},
    "restriction_08_FoodCoop": {"offset": 89, "length": 1},
    "restriction_09_BCNeomycin": {"offset": 90, "length": 1},
    "restriction_10_ABLSDisease": {"offset": 91, "length": 1},
    "restriction_11_FrozenFood": {"offset": 92, "length": 1},
    "restriction_12_CropSupp": {"offset": 93, "length": 1},
    "restriction_13_WestBest": {"offset": 94, "length": 1},
    "restriction_14_ONTobTax": {"offset": 95, "length": 1},
    "restriction_15_FeedRestr": {"offset": 96, "length": 1},
    "restriction_16_HarmonCtry": {"offset": 97, "length": 1},
    "restriction_17_MNHBC": {"offset": 98, "length": 1},
    "restriction_18_ABHBC": {"offset": 99, "length": 1},
    "restriction_19_NTHBC": {"offset": 100, "length": 1},
    "restriction_20_CropSupp": {"offset": 101, "length": 1},
    "restriction_21_MNTobTax": {"offset": 102, "length": 1},
    "restriction_22_SKTobTax": {"offset": 103, "length": 1},
    "restriction_23_ABTobTax": {"offset": 104, "length": 1},
    "restriction_24_NTTobTax": {"offset": 105, "length": 1},
    "restriction_25_HdwNonMember": {"offset": 106, "length": 1},
    "restriction_26_ABPest": {"offset": 107, "length": 1},
    "restriction_27_HdwSeasonal": {"offset": 108, "length": 1},
    "restriction_28_ArcticCoop": {"offset": 109, "length": 1},
    "restriction_29_Smittys": {"offset": 110, "length": 1},
    "restriction_30_PestSched1": {"offset": 111, "length": 1},
    "restriction_31_PestSched2": {"offset": 112, "length": 1},
    "restriction_32_PestSched3": {"offset": 113, "length": 1},
    "restriction_33_NUTobTax": {"offset": 114, "length": 1},
    "restriction_34_CashCarry": {"offset": 115, "length": 1},
    "restriction_35_CHA": {"offset": 116, "length": 1},
    "restriction_36_BCGMRestr": {"offset": 117, "length": 1},
    "reserved_for_restriction_37": {"offset": 118, "length": 1},
    "reserved_for_restriction_38": {"offset": 119, "length": 1},
    "reserved_for_restriction_39": {"offset": 120, "length": 1},
    "reserved_for_restriction_40": {"offset": 121, "length": 1},
    "large_bottle_deposite": {"offset": 122, "length": 6},
    "brand_code": {"offset": 128, "length": 1},
    "EDBV_item": {"offset": 129, "length": 1},
    "annual_sales": {"offset": 130, "length": 7},
    "last_year_annual_sales": {"offset": 137, "length": 7},
    "layer_quantity": {"offset": 144, "length": 4},
    "pallet_quantity": {"offset": 148, "length": 4},
    "minimum_order_quantity": {"offset": 152, "length": 4},
    "major_department": {"offset": 156, "length": 2},
    "minor_department": {"offset": 158, "length": 3},
    "major_category": {"offset": 161, "length": 2},
    "minor_category": {"offset": 163, "length": 3},
    "master_item_number": {"offset": 166, "length": 10},
    "substitute_item_number": {"offset": 176, "length": 10},
    "quantity_in_warehouse": {"offset": 186, "length": 8},
    "ecology_code": {"offset": 194, "length": 8},
    "reclamation_status": {"offset": 202, "length": 1},
}

file_layout_dict_b02 = {
    "type": {"offset": 0, "length": 1},
    "code": {"offset": 1, "length": 2},
    "product_number": {"offset": 3, "length": 10},
    "store_vendor_id": {"offset": 13, "length": 10},
    "group_number": {"offset": 23, "length": 2},
    "family_number": {"offset": 25, "length": 6},
    "scan_number": {"offset": 31, "length": 20},
    "case_upc": {"offset": 51, "length": 20},
    "store_location": {"offset": 71, "length": 6},
    "aisle_location": {"offset": 71, "length": 2},
    "side_location": {"offset": 73, "length": 1},
    "shelf_location": {"offset": 74, "length": 3},
    "department_number": {"offset": 77, "length": 2},
    "alternate_upc_1": {"offset": 79, "length": 20},
    "alternate_upc_2": {"offset": 99, "length": 20},
    "alternate_upc_3": {"offset": 119, "length": 20},
    "alternate_upc_4": {"offset": 139, "length": 20},
    "alternate_upc_5": {"offset": 159, "length": 20},
    "fcl_vendor_id": {"offset": 179, "length": 10},
    "originating_warehouse": {"offset": 189, "length": 1},
    "gluten_free_code": {"offset": 190, "length": 1},
}

file_layout_dict_p01 = {
    "type": {"offset": 0, "length": 1},
    "code": {"offset": 1, "length": 2},
    "product_number": {"offset": 3, "length": 10},
    "retail_for": {"offset": 13, "length": 2},
    "retail_price": {"offset": 15, "length": 6},
    "case_cost":    {"offset": 21, "length": 6},
    "cube_length": {"offset": 27, "length": 6},
    "cube_width": {"offset": 33, "length": 6},
    "cube_height": {"offset": 39, "length": 6},
    "cube_weight_lbs": {"offset": 45, "length": 6},
    "shelf_height": {"offset": 51, "length": 6},
    "shelf_width": {"offset": 57, "length": 6},
    "shelf_depth": {"offset": 63, "length": 6},
    "ecology_container_fee_case": {"offset": 69, "length": 4},
    "sequence_within_family": {"offset": 73, "length": 5},
    "cost_includes_ecology": {"offset": 78, "length": 1},
    "estimated_freight": {"offset": 79, "length": 6},
    "EDBV_adjustment": {"offset": 85, "length": 5}, # In the doc it says "999v99" not sure what the means.
    "EDBV_adjustment_sign": {"offset": 90, "length": 1},
    "cost_adjustment": {"offset": 91, "length": 5}, # In the doc it says "999v99" not sure what the means.
    "cost_adjustment_sign": {"offset": 96, "length": 1},
    "large_ecology": {"offset": 97, "length": 6},
    "store_target_margin": {"offset": 103, "length": 6},
    "SRP_target_margin": {"offset": 109, "length": 6},
    "SRP_price": {"offset": 115, "length": 6},
    "tobacco_tax": {"offset": 121, "length": 6},
    "case_cost_new": {"offset": 127, "length": 6},
    "SCC_code_1": {"offset": 133, "length": 20},
    "SCC_code_2": {"offset": 153, "length": 20},
    "SCC_code_3": {"offset": 173, "length": 20},
    "SCC_code_4": {"offset": 193, "length": 20},
    "SCC_code_5": {"offset": 213, "length": 20}
}

# Note: V01, V02, etc. isn't really used. But if needed I can add it.
# Example: Getting a specific entry using the dictionary. 
# ==> print(product_line[file_layout_dict_b01["product_number"]["offset"]:file_layout_dict_b01["product_number"]["offset"] + file_layout_dict_b01["product_number"]["length"]])

def retrieve_specific_entry(line, layout_dict, entry_name):
    if entry_name not in layout_dict:
        raise ValueError(f"Entry name '{entry_name}' not found in layout dictionary.")
    
    offset = layout_dict[entry_name]["offset"]
    length = layout_dict[entry_name]["length"]
    
    return line[offset:offset + length].strip()

def retrieve_all_b01_entries(line):
    entries = {}
    for entry_name in file_layout_dict_b01:
        entries[entry_name] = retrieve_specific_entry(line, file_layout_dict_b01, entry_name)
    return entries

def retrieve_all_b02_entries(line):
    entries = {}
    for entry_name in file_layout_dict_b02:
        entries[entry_name] = retrieve_specific_entry(line, file_layout_dict_b02, entry_name)
    return entries

def retrieve_all_p01_entries(line):
    entries = {}
    for entry_name in file_layout_dict_p01:
        entries[entry_name] = retrieve_specific_entry(line, file_layout_dict_p01, entry_name)
    return entries

def export_to_excel(products, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Catalog Master"
    
    # Ensure standard gridlines are active
    ws.views.sheetView[0].showGridLines = True
    
    # --- Cohesive Design Palettes (Soft Pastels) ---
    # Top Merged Header: Dark Slate/Navy for crisp contrast
    navy_header_fill = PatternFill(start_color="2F3E46", end_color="2F3E46", fill_type="solid")
    font_main_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    # Section 1 (B01): Soft Pastel Blue
    b01_sub_fill = PatternFill(start_color="DCEEFF", end_color="DCEEFF", fill_type="solid") # Stronger pastel for header
    b01_row_fill_a = PatternFill(start_color="F2F8FF", end_color="F2F8FF", fill_type="solid") # Main data row
    b01_row_fill_b = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid") # Zebra alternative
    
    # Section 2 (B02): Soft Pastel Green
    b02_sub_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid") 
    b02_row_fill_a = PatternFill(start_color="F4F9F1", end_color="F4F9F1", fill_type="solid") 
    b02_row_fill_b = PatternFill(start_color="EBF5E6", end_color="EBF5E6", fill_type="solid") 
    
    # Section 3 (P01): Soft Pastel Purple/Lavender
    p01_sub_fill = PatternFill(start_color="E8E1F5", end_color="E8E1F5", fill_type="solid") 
    p01_row_fill_a = PatternFill(start_color="F6F3FA", end_color="F6F3FA", fill_type="solid") 
    p01_row_fill_b = PatternFill(start_color="EFEAF6", end_color="EFEAF6", fill_type="solid") 

    # Fonts and Borders
    font_sub_header = Font(name="Segoe UI", size=10, bold=True, color="333333")
    font_data = Font(name="Segoe UI", size=10, color="222222")
    thin_border_side = Side(style='thin', color='D9D9D9') # Slightly softer than E0E0E0 for pastels
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # 1. Generate Merged Top-Level Segment Headers
    segments = [
        ("Core Item Demographics (B01)", len(file_layout_dict_b01)),
        ("Logistics & Warehouse Paths (B02)", len(file_layout_dict_b02)),
        ("Pricing & Financial Metrics (P01)", len(file_layout_dict_p01))
    ]
    
    current_col = 1
    for name, length in segments:
        ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + length - 1)
        header_cell = ws.cell(row=1, column=current_col, value=name)
        header_cell.font = font_main_header
        header_cell.fill = navy_header_fill
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        current_col += length
    ws.row_dimensions[1].height = 26

    # 2. Write Lower Field Level Sub-Headers with Pastel Backgrounds
    all_fields = (
        list(file_layout_dict_b01.keys()) + 
        list(file_layout_dict_b02.keys()) + 
        list(file_layout_dict_p01.keys())
    )
    
    len_b01 = len(file_layout_dict_b01)
    len_b02 = len(file_layout_dict_b02)
    
    for col_idx, field_key in enumerate(all_fields, start=1):
        clean_title = field_key.replace("_", " ").title()
        sub_cell = ws.cell(row=2, column=col_idx, value=clean_title)
        sub_cell.font = font_sub_header
        sub_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        sub_cell.border = cell_border
        
        # Color code sub-headers based on section boundaries
        if col_idx <= len_b01:
            sub_cell.fill = b01_sub_fill
        elif col_idx <= (len_b01 + len_b02):
            sub_cell.fill = b02_sub_fill
        else:
            sub_cell.fill = p01_sub_fill
            
    ws.row_dimensions[2].height = 28
    
    # 3. Populate Rows
    current_row = 3
    for product in products:
        prod_key = list(product.keys())[0]
        data_map = product[prod_key]
        
        # Flatten all values sequentially matching column ordering
        row_values = []
        for f in file_layout_dict_b01.keys(): row_values.append(data_map["B01_entries"].get(f, ""))
        for f in file_layout_dict_b02.keys(): row_values.append(data_map["B02_entries"].get(f, ""))
        for f in file_layout_dict_p01.keys(): row_values.append(data_map["P01_entries"].get(f, ""))
        
        # Determine if this row is an alternating row for zebra striping
        is_zebra_row = (current_row % 2 == 0)
        
        # Write values safely and apply section color coding
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_data
            cell.border = cell_border
            
            # Color code data cells + apply pastel zebra striping
            if col_idx <= len_b01:
                cell.fill = b01_row_fill_b if is_zebra_row else b01_row_fill_a
            elif col_idx <= (len_b01 + len_b02):
                cell.fill = b02_row_fill_b if is_zebra_row else b02_row_fill_a
            else:
                cell.fill = p01_row_fill_b if is_zebra_row else p01_row_fill_a
                
            # Formatting numbers safely
            if isinstance(val, float):
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(val, int):
                cell.alignment = Alignment(horizontal="right")
                
        current_row += 1
        
    # 4. Sheet Polish: Set Panes and Auto-Width Fit Columns
    ws.freeze_panes = "D3" # Freezes row 1 & 2 headers, and columns A-C
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
        
    wb.save(filename)
    print(f"Spreadsheet generated successfully: {filename}")

def main():
    products = []
    current_product = {}
    
    with open(BK1_FILEPATH, "r") as f: 
        for line in f:
            line = line.rstrip('\n') # Removes only newline chars.
            if not line:
                continue
            
            # Identify record type (B01, B02, P01)
            record_type = line[:3]
            
            # Start a new product group (Note: Each item gets 3 lines associated with it B01, B02 P01. Thus, if there are 15 items then there'll be 15 x 3 = 45 lines in the file).
            if record_type == "B01":
                product_name = line[file_layout_dict_b01["product_description"]["offset"]:file_layout_dict_b01["product_description"]["offset"] + file_layout_dict_b01["product_description"]["length"]].strip()

                current_product = {
                    product_name: {
                        "B01": line,
                        "B02": None,
                        "P01": None,
                    }
                }

            elif record_type == "B02":
                if current_product:
                    # Get the active product name key and inject the B02 line
                    name = list(current_product.keys())[0]
                    current_product[name]["B02"] = line

            elif record_type == "P01":
                if current_product:
                    # Get the active product name key and inject the P01 line
                    name = list(current_product.keys())[0]
                    current_product[name]["P01"] = line

                    # P01 signals the end of the 3-line block, save it to our list
                    products.append(current_product)
                    current_product = {}  # Reset for the next product
    
    # For each product, we can now parse the B01, B02, and P01 lines into their respective entries using the layout dictionaries.
    for product in products:
        name = list(product.keys())[0]
        product[name]["B01_entries"] = retrieve_all_b01_entries(product[name]["B01"])
        product[name]["B02_entries"] = retrieve_all_b02_entries(product[name]["B02"])
        product[name]["P01_entries"] = retrieve_all_p01_entries(product[name]["P01"]) 
    
    # Finally -> Export the products list as excel.
    export_to_excel(products, EXCEL_OUTPUT_PATH)
    
    
    print(json.dumps(products, indent=4))

main()